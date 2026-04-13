"""
compare_excel.py
────────────────
Compares two Excel workbooks found in the same folder and writes a
human + LLM readable Markdown report: sheet_diff.md

Detects:
  • Sheets added / removed / renamed (by similarity)
  • Column additions, removals, renames, and moves
  • Row count changes
  • Cell value changes (columns aligned by name, not position)
  • Formula changes
  • Chart additions / removals
  • Sort-order changes (same data, different row order)
"""

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
from difflib import SequenceMatcher
from datetime import datetime


# ── Config ───────────────────────────────────────────────────────────────────

# Max changed cells to list individually per column (avoids huge output)
MAX_CELL_CHANGES_SHOWN = 10

# Similarity threshold (0–1) for fuzzy sheet-name matching
SHEET_NAME_SIMILARITY = 0.6


# ── Helpers ──────────────────────────────────────────────────────────────────

def find_table_start(df_raw: pd.DataFrame, min_fill_ratio: float = 0.5) -> int:
    for i, row in df_raw.iterrows():
        non_null = row.notna().sum()
        total = len(row)
        if non_null >= 2 and (non_null / total) >= min_fill_ratio:
            return i
    return 0


def load_sheet(xl_pd: pd.ExcelFile, wb: openpyxl.Workbook, sheet: str):
    """
    Returns (df, header_row_idx, formula_dict, chart_count) for a sheet.
    df has clean column names and empty rows/cols removed.
    formula_dict: {col_name: sample_formula}
    """
    ws = wb[sheet]

    # Collect formulas
    formula_map = {}
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_map[(cell.row - 1, cell.column - 1)] = cell.value

    # Chart count
    chart_count = len(ws._charts) if hasattr(ws, "_charts") else 0

    # Raw parse to detect header
    df_raw = xl_pd.parse(sheet, header=None)
    if df_raw.empty:
        return pd.DataFrame(), 0, {}, chart_count

    header_row_idx = find_table_start(df_raw)
    df_full = xl_pd.parse(sheet, header=header_row_idx)

    surviving_mask = ~df_full.isna().all(axis=0)
    surviving_positions = [i for i, keep in enumerate(surviving_mask) if keep]
    df = df_full.loc[:, surviving_mask].dropna(axis=0, how="all").reset_index(drop=True)

    df.columns = [
        f"Unnamed_col_{i}" if str(c).startswith("Unnamed:") else str(c)
        for i, c in enumerate(df.columns)
    ]

    # Map col_name → sample formula
    data_start_row = header_row_idx + 1
    col_formulas = {}
    for col_name, orig_i in zip(df.columns, surviving_positions):
        samples = [
            f for (r, c), f in formula_map.items()
            if c == orig_i and r >= data_start_row
        ]
        if samples:
            col_formulas[col_name] = samples[0]

    return df, header_row_idx, col_formulas, chart_count


def fuzzy_score(a: str, b: str) -> float:
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()


def match_sheets(sheets_a: list, sheets_b: list) -> dict:
    """
    Returns {sheet_a_name: sheet_b_name | None}.
    Exact matches first, then fuzzy, then unmatched → None.
    """
    matched = {}
    used_b = set()

    # Exact matches
    for s in sheets_a:
        if s in sheets_b:
            matched[s] = s
            used_b.add(s)

    # Fuzzy matches for remaining
    for s in sheets_a:
        if s in matched:
            continue
        candidates = [b for b in sheets_b if b not in used_b]
        if not candidates:
            matched[s] = None
            continue
        best = max(candidates, key=lambda b: fuzzy_score(s, b))
        score = fuzzy_score(s, best)
        if score >= SHEET_NAME_SIMILARITY:
            matched[s] = best
            used_b.add(best)
        else:
            matched[s] = None

    return matched


def detect_column_moves(cols_a: list, cols_b: list) -> list[tuple]:
    """
    For columns present in both, return list of (col, old_idx, new_idx)
    where position changed.
    """
    common = [c for c in cols_a if c in cols_b]
    moves = []
    for col in common:
        pos_a = cols_a.index(col)
        pos_b = cols_b.index(col)
        if pos_a != pos_b:
            moves.append((col, pos_a + 1, pos_b + 1))  # 1-based for readability
    return moves


def detect_sort_change(df_a: pd.DataFrame, df_b: pd.DataFrame, common_cols: list) -> str | None:
    """
    If both frames have identical sets of rows but different order,
    return a description. Uses a frozenset of row tuples.
    """
    if df_a.shape != df_b.shape:
        return None
    try:
        subset = [c for c in common_cols if c in df_a.columns and c in df_b.columns]
        if not subset:
            return None
        set_a = set(map(tuple, df_a[subset].astype(str).values.tolist()))
        set_b = set(map(tuple, df_b[subset].astype(str).values.tolist()))
        if set_a == set_b:
            # Same rows, check if order differs
            if df_a[subset].astype(str).values.tolist() != df_b[subset].astype(str).values.tolist():
                return "Row order changed (same data, different sort)"
    except Exception:
        pass
    return None


def compare_values(df_a: pd.DataFrame, df_b: pd.DataFrame, common_cols: list) -> list[str]:
    """
    Compare cell values column by column for common columns.
    Returns list of finding strings.
    """
    findings = []
    min_rows = min(len(df_a), len(df_b))

    def normalize(series: pd.Series) -> pd.Series:
        """
        Blank out anything that was originally null — checked on the raw series
        BEFORE astype(str), which avoids unreliable string matching of 'nan'.
        """
        sliced   = series.iloc[:min_rows].reset_index(drop=True)
        is_empty = sliced.isna()
        as_str   = sliced.astype(str).str.strip()
        null_words = {"nan", "none", "nat", ""}
        as_str[is_empty | as_str.str.lower().isin(null_words)] = ""
        return as_str

    for col in common_cols:
        if col not in df_a.columns or col not in df_b.columns:
            continue
        col_a = normalize(df_a[col])
        col_b = normalize(df_b[col])
        # Exclude rows where both sides are empty (false positives from NaN variants)
        both_empty = (col_a == "") & (col_b == "")
        diff_mask = (col_a != col_b) & ~both_empty
        diff_count = diff_mask.sum()

        if diff_count == 0:
            continue

        pct = diff_count / min_rows * 100
        changes = []
        for row_i in diff_mask[diff_mask].index[:MAX_CELL_CHANGES_SHOWN]:
            val_a = col_a[row_i]
            val_b = col_b[row_i]
            # Skip if both are empty after normalization (was a false positive)
            if val_a == "" and val_b == "":
                continue
            excel_row = row_i + 2  # +1 for header, +1 for 1-based
            changes.append(f"row {excel_row}: `{val_a}` → `{val_b}`")

        more = f" _(+{diff_count - MAX_CELL_CHANGES_SHOWN} more)_" if diff_count > MAX_CELL_CHANGES_SHOWN else ""
        findings.append(
            f"  - **{col}**: {diff_count} cell(s) changed ({pct:.1f}% of compared rows)\n"
            + "\n".join(f"    - {c}" for c in changes)
            + more
        )

    return findings


def compare_formulas(formulas_a: dict, formulas_b: dict) -> list[str]:
    """Compare formula samples per column."""
    findings = []
    all_cols = set(formulas_a) | set(formulas_b)
    for col in sorted(all_cols):
        fa = formulas_a.get(col)
        fb = formulas_b.get(col)
        if fa and not fb:
            findings.append(f"  - **{col}**: formula removed (was `{fa}`)")
        elif fb and not fa:
            findings.append(f"  - **{col}**: formula added (`{fb}`)")
        elif fa and fb and fa != fb:
            findings.append(f"  - **{col}**: formula changed\n    - Before: `{fa}`\n    - After:  `{fb}`")
    return findings


# ── Sheet comparison ──────────────────────────────────────────────────────────

def compare_sheet_pair(
    sheet_a: str, sheet_b: str,
    xl_a: pd.ExcelFile, wb_a: openpyxl.Workbook,
    xl_b: pd.ExcelFile, wb_b: openpyxl.Workbook,
) -> str:
    lines = []

    renamed = sheet_a != sheet_b
    if renamed:
        lines.append(f"> ⚠ Sheet was **renamed**: `{sheet_a}` → `{sheet_b}`\n")

    df_a, hdr_a, formulas_a, charts_a = load_sheet(xl_a, wb_a, sheet_a)
    df_b, hdr_b, formulas_b, charts_b = load_sheet(xl_b, wb_b, sheet_b)

    # ── Empty sheet ──
    if df_a.empty and df_b.empty:
        lines.append("_Both versions of this sheet are empty._")
        return "\n".join(lines)
    if df_a.empty:
        lines.append("_Sheet was **empty in File A** but has data in File B._")
        return "\n".join(lines)
    if df_b.empty:
        lines.append("_Sheet had data in File A but is **empty in File B**._")
        return "\n".join(lines)

    # ── Table start ──
    if hdr_a != hdr_b:
        lines.append(f"- Table header row shifted: row **{hdr_a + 1}** → row **{hdr_b + 1}** "
                     f"(leading note rows changed)\n")

    # ── Shape ──
    rows_a, cols_a_n = df_a.shape
    rows_b, cols_b_n = df_b.shape
    lines.append(f"### Shape\n")
    lines.append(f"| | File A | File B | Δ |")
    lines.append(f"|---|---|---|---|")
    lines.append(f"| Rows | {rows_a} | {rows_b} | {rows_b - rows_a:+d} |")
    lines.append(f"| Columns | {cols_a_n} | {cols_b_n} | {cols_b_n - cols_a_n:+d} |\n")

    cols_a = list(df_a.columns)
    cols_b = list(df_b.columns)
    set_a = set(cols_a)
    set_b = set(cols_b)

    # ── Column additions / removals ──
    added_cols   = sorted(set_b - set_a)
    removed_cols = sorted(set_a - set_b)
    common_cols  = [c for c in cols_a if c in set_b]

    if added_cols or removed_cols:
        lines.append("### Column Changes\n")
        for c in added_cols:
            pos = cols_b.index(c) + 1
            col_letter = get_column_letter(pos)
            lines.append(f"- ✅ **Added**: `{c}` (col {col_letter}, position {pos})")
        for c in removed_cols:
            pos = cols_a.index(c) + 1
            col_letter = get_column_letter(pos)
            lines.append(f"- ❌ **Removed**: `{c}` (was col {col_letter}, position {pos})")
        lines.append("")

    # ── Column moves ──
    moves = detect_column_moves(cols_a, cols_b)
    if moves:
        lines.append("### Column Moves\n")
        for col, old_pos, new_pos in moves:
            lines.append(f"- **{col}**: position {old_pos} ({get_column_letter(old_pos)}) "
                         f"→ position {new_pos} ({get_column_letter(new_pos)})")
        lines.append("")

    # ── Sort change ──
    sort_note = detect_sort_change(df_a, df_b, common_cols)
    if sort_note:
        lines.append(f"### Row Order\n\n- ℹ️ {sort_note}\n")

    # ── Value changes ──
    value_findings = compare_values(df_a, df_b, common_cols)
    if value_findings:
        lines.append("### Value Changes\n")
        lines.extend(value_findings)
        lines.append("")

    # ── Formula changes ──
    formula_findings = compare_formulas(formulas_a, formulas_b)
    if formula_findings:
        lines.append("### Formula Changes\n")
        lines.extend(formula_findings)
        lines.append("")

    # ── Charts ──
    if charts_a != charts_b:
        lines.append("### Charts\n")
        if charts_b > charts_a:
            lines.append(f"- 📊 {charts_b - charts_a} chart(s) **added** "
                         f"({charts_a} → {charts_b})")
        else:
            lines.append(f"- 📊 {charts_a - charts_b} chart(s) **removed** "
                         f"({charts_a} → {charts_b})")
        lines.append("")

    if len(lines) == 0 or all(l.startswith("|") or l == "" for l in lines):
        lines.append("_No differences detected in this sheet._")

    return "\n".join(lines)


# ── Main ─────────────────────────────────────────────────────────────────────

script_dir = Path(__file__).parent.absolute()
excel_files = list(script_dir.glob("*.xlsx")) + list(script_dir.glob("*.xls"))

if len(excel_files) < 2:
    print(f"❌ Need exactly 2 Excel files in {script_dir}, found {len(excel_files)}.")
    exit(1)

if len(excel_files) > 2:
    print(f"⚠ Found {len(excel_files)} Excel files. Using the two oldest by modification time.")
    excel_files = sorted(excel_files, key=lambda p: p.stat().st_mtime)[:2]

file_a, file_b = excel_files[0], excel_files[1]
print(f"Comparing:\n  A: {file_a.name}\n  B: {file_b.name}\n")

try:
    xl_a  = pd.ExcelFile(file_a)
    xl_b  = pd.ExcelFile(file_b)
    wb_a  = openpyxl.load_workbook(file_a, data_only=False)
    wb_b  = openpyxl.load_workbook(file_b, data_only=False)
except Exception as e:
    print(f"❌ Failed to open files: {e}")
    exit(1)

sheets_a = xl_a.sheet_names
sheets_b = xl_b.sheet_names
sheet_match = match_sheets(sheets_a, sheets_b)
only_in_b = [s for s in sheets_b if s not in sheet_match.values()]

output_path = script_dir / "sheet_diff.md"

with open(output_path, "w", encoding="utf-8") as f:

    # ── Header ──
    f.write(f"# Excel Comparison Report\n\n")
    f.write(f"| | |  \n")
    f.write(f"|---|---|\n")
    f.write(f"| **File A** | `{file_a.name}` |\n")
    f.write(f"| **File B** | `{file_b.name}` |\n")
    f.write(f"| **Generated** | {datetime.now().strftime('%Y-%m-%d %H:%M')} |\n\n")

    # ── Sheet overview ──
    f.write(f"## Sheet Overview\n\n")
    f.write(f"| Sheet (A) | Sheet (B) | Status |\n")
    f.write(f"|---|---|---|\n")
    for sa, sb in sheet_match.items():
        if sb is None:
            f.write(f"| `{sa}` | — | ❌ Removed |\n")
        elif sa == sb:
            f.write(f"| `{sa}` | `{sb}` | ✅ Matched |\n")
        else:
            f.write(f"| `{sa}` | `{sb}` | ✏️ Renamed |\n")
    for sb in only_in_b:
        f.write(f"| — | `{sb}` | ✅ Added |\n")
    f.write("\n")

    # ── Per-sheet diff ──
    f.write("---\n\n")
    for sa, sb in sheet_match.items():
        if sb is None:
            f.write(f"## Sheet: `{sa}`\n\n")
            f.write(f"❌ **This sheet was removed in File B.**\n\n---\n\n")
            continue

        display_name = sb if sa == sb else f"{sa} → {sb}"
        f.write(f"## Sheet: `{display_name}`\n\n")
        section = compare_sheet_pair(sa, sb, xl_a, wb_a, xl_b, wb_b)
        f.write(section + "\n\n---\n\n")

    for sb in only_in_b:
        f.write(f"## Sheet: `{sb}`\n\n")
        f.write(f"✅ **This sheet was added in File B.**\n\n---\n\n")

print(f"✓ Report written to: {output_path.name}")