from __future__ import annotations

import posixpath
import tempfile
import xml.etree.ElementTree as ET
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from xml.sax.saxutils import escape

from .constants import XML_NS

DRAWING_REL_TYPE = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing"
DRAWING_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.drawing+xml"


def _new_workbook() -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)
    return workbook


def _save(workbook: Workbook, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def create_simple_table_workbook(path: Path) -> Path:
    workbook = _new_workbook()
    ws = workbook.create_sheet("Simple")
    ws.append(["Name", "Score", "Active"])
    ws.append(["Alice", 10, True])
    ws.append(["Bob", 12, False])
    ws.append(["Charlie", 9, True])
    return _save(workbook, path)


def create_metadata_gap_workbook(path: Path) -> Path:
    workbook = _new_workbook()
    ws = workbook.create_sheet("GapHeader")
    ws.append(["Monthly KPI Export", None, None, None])
    ws.append([None, None, None, None])
    ws.append(["Region", "Sales", "Margin", "Date"])
    ws.append(["North", 120, 0.32, "2026-03-31"])
    ws.append(["South", 95, 0.28, "2026-03-31"])
    return _save(workbook, path)


def create_manual_override_workbook(path: Path) -> Path:
    workbook = _new_workbook()
    ws = workbook.create_sheet("ManualHeader")
    ws.append(["Staffy", "Staffy", "Staffy", "Staffy", "Formullu"])
    ws.append([None, None, None, "31.03.2026", "458.025,00"])
    ws.append(["Marka", "Sicil No", "Adi Soyadi", "Kidem", "Nihai Prim"])
    ws.append(["M&S", 25586, "Berivan Ozturk", 1.70, 0])
    ws.append(["M&S", 22007, "Deniz Picak", 6.75, 3000])
    return _save(workbook, path)


def create_pivot_like_workbook(path: Path) -> Path:
    workbook = _new_workbook()
    ws = workbook.create_sheet("PivotLike")
    ws.append(["Department", "(Multiple Items)", None])
    ws.append([None, None, None])
    ws.append(["Row Labels", "Sum of Sales", "Count of Orders"])
    ws.append(["Shoes", 1500, 12])
    ws.append(["Kids", 850, 7])
    return _save(workbook, path)


def create_formula_workbook(path: Path) -> Path:
    workbook = _new_workbook()
    ws = workbook.create_sheet("Formulas")
    ws.append(["Qty", "Unit Price", "Total"])
    ws.append([2, 5, "=A2*B2"])
    ws.append([4, 7, "=A3*B3"])
    ws.append([1, 9, "=A4*B4"])
    return _save(workbook, path)


def create_unnamed_header_workbook(path: Path) -> Path:
    workbook = _new_workbook()
    ws = workbook.create_sheet("Unnamed")
    ws.append(["Employee", None, "Store"])
    ws.append(["Ada", "FT", "Istanbul"])
    ws.append(["Can", "PT", "Ankara"])
    return _save(workbook, path)


def create_empty_and_sparse_workbook(path: Path) -> Path:
    workbook = _new_workbook()
    workbook.create_sheet("Empty")
    ws = workbook.create_sheet("Sparse")
    ws["A4"] = "Header A"
    ws["B4"] = "Header B"
    ws["A5"] = "Value 1"
    ws["B5"] = 10
    return _save(workbook, path)


def create_multi_sheet_workbook(path: Path) -> Path:
    workbook = _new_workbook()

    simple = workbook.create_sheet("Overview")
    simple.append(["Metric", "Value"])
    simple.append(["Stores", 12])
    simple.append(["Employees", 315])

    notes = workbook.create_sheet("NotesBeforeTable")
    notes.append(["Prepared by Finance", None, None])
    notes.append([None, None, None])
    notes.append(["Month", "Amount", "Approved"])
    notes.append(["2026-01", 100, "Yes"])
    notes.append(["2026-02", 120, "No"])

    formulas = workbook.create_sheet("FormulaSheet")
    formulas.append(["Base", "Rate", "Result"])
    formulas.append([10, 0.2, "=A2*B2"])
    formulas.append([20, 0.3, "=A3*B3"])

    return _save(workbook, path)


def create_textbox_workbook(path: Path) -> Path:
    workbook = _new_workbook()
    ws = workbook.create_sheet("TextboxSheet")
    ws.append(["Name", "Decision"])
    ws.append(["Aylin", "Review"])
    ws.append(["Mert", "Approved"])
    _save(workbook, path)
    inject_textbox(path, "TextboxSheet", "Floating business rule\nSecond line", "A5")
    return path


def build_all_test_workbooks(output_dir: Path) -> list[Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    workbook_paths = [
        create_simple_table_workbook(output_dir / "simple_table.xlsx"),
        create_metadata_gap_workbook(output_dir / "metadata_gap.xlsx"),
        create_manual_override_workbook(output_dir / "manual_override.xlsx"),
        create_pivot_like_workbook(output_dir / "pivot_like.xlsx"),
        create_formula_workbook(output_dir / "formula_table.xlsx"),
        create_unnamed_header_workbook(output_dir / "unnamed_header.xlsx"),
        create_empty_and_sparse_workbook(output_dir / "empty_and_sparse.xlsx"),
        create_multi_sheet_workbook(output_dir / "multi_sheet_suite.xlsx"),
        create_textbox_workbook(output_dir / "textbox_sheet.xlsx"),
    ]
    return workbook_paths


def _next_relationship_id(root: ET.Element) -> str:
    existing_ids: set[int] = set()
    for rel in root.findall(f"{{{XML_NS['pkgrel']}}}Relationship"):
        rel_id = rel.attrib.get("Id", "")
        if rel_id.startswith("rId") and rel_id[3:].isdigit():
            existing_ids.add(int(rel_id[3:]))

    next_id = 1
    while next_id in existing_ids:
        next_id += 1
    return f"rId{next_id}"


def _sheet_part_for_name(parts: dict[str, bytes], sheet_name: str) -> str:
    workbook_root = ET.fromstring(parts["xl/workbook.xml"])
    workbook_rels = ET.fromstring(parts["xl/_rels/workbook.xml.rels"])
    rel_map = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in workbook_rels.findall("pkgrel:Relationship", XML_NS)
    }

    for sheet in workbook_root.find("main:sheets", XML_NS):
        if sheet.attrib.get("name") != sheet_name:
            continue
        rel_id = sheet.attrib.get(f"{{{XML_NS['office_rel']}}}id")
        if rel_id and rel_id in rel_map:
            target = rel_map[rel_id]
            return posixpath.normpath(posixpath.join("xl", target)).lstrip("/")

    raise KeyError(f"Sheet '{sheet_name}' not found in workbook package.")


def _anchor_parts(anchor: str) -> tuple[int, int]:
    letters = "".join(ch for ch in anchor if ch.isalpha())
    numbers = "".join(ch for ch in anchor if ch.isdigit())
    if not letters or not numbers:
        raise ValueError(f"Invalid anchor '{anchor}'.")

    col_idx = 0
    for char in letters.upper():
        col_idx = col_idx * 26 + (ord(char) - ord("A") + 1)
    return col_idx - 1, int(numbers) - 1


def _drawing_xml(text: str, anchor: str, shape_name: str) -> bytes:
    col_idx, row_idx = _anchor_parts(anchor)
    end_col = col_idx + 4
    end_row = row_idx + 3
    paragraphs = "".join(
        f"<a:p><a:r><a:rPr lang=\"en-US\" sz=\"1100\"/><a:t>{escape(line)}</a:t></a:r></a:p>"
        for line in text.splitlines()
    )
    xml = (
        "<?xml version=\"1.0\" encoding=\"UTF-8\" standalone=\"yes\"?>"
        "<xdr:wsDr "
        "xmlns:xdr=\"http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing\" "
        "xmlns:a=\"http://schemas.openxmlformats.org/drawingml/2006/main\">"
        "<xdr:twoCellAnchor>"
        f"<xdr:from><xdr:col>{col_idx}</xdr:col><xdr:colOff>0</xdr:colOff>"
        f"<xdr:row>{row_idx}</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>"
        f"<xdr:to><xdr:col>{end_col}</xdr:col><xdr:colOff>0</xdr:colOff>"
        f"<xdr:row>{end_row}</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:to>"
        "<xdr:sp macro=\"\" textlink=\"\">"
        "<xdr:nvSpPr>"
        f"<xdr:cNvPr id=\"2\" name=\"{escape(shape_name)}\"/>"
        "<xdr:cNvSpPr txBox=\"1\"/>"
        "</xdr:nvSpPr>"
        "<xdr:spPr><a:prstGeom prst=\"rect\"><a:avLst/></a:prstGeom></xdr:spPr>"
        "<xdr:txBody><a:bodyPr/><a:lstStyle/>"
        f"{paragraphs}</xdr:txBody>"
        "</xdr:sp><xdr:clientData/></xdr:twoCellAnchor></xdr:wsDr>"
    )
    return xml.encode("utf-8")


def inject_textbox(path: Path, sheet_name: str, text: str, anchor: str, shape_name: str = "TextBox 1") -> None:
    with ZipFile(path, "r") as zip_file:
        parts = {name: zip_file.read(name) for name in zip_file.namelist()}

    sheet_part = _sheet_part_for_name(parts, sheet_name)
    sheet_file_name = posixpath.basename(sheet_part)
    sheet_rels_path = f"xl/worksheets/_rels/{sheet_file_name}.rels"

    if sheet_rels_path in parts:
        rels_root = ET.fromstring(parts[sheet_rels_path])
    else:
        rels_root = ET.Element(f"{{{XML_NS['pkgrel']}}}Relationships")

    existing_drawing_targets = {
        rel.attrib.get("Target", "")
        for rel in rels_root.findall(f"{{{XML_NS['pkgrel']}}}Relationship")
        if rel.attrib.get("Type") == DRAWING_REL_TYPE
    }
    next_drawing_idx = 1
    while f"../drawings/drawing{next_drawing_idx}.xml" in existing_drawing_targets:
        next_drawing_idx += 1

    drawing_rel_target = f"../drawings/drawing{next_drawing_idx}.xml"
    drawing_part = f"xl/drawings/drawing{next_drawing_idx}.xml"
    relationship_id = _next_relationship_id(rels_root)

    ET.SubElement(
        rels_root,
        f"{{{XML_NS['pkgrel']}}}Relationship",
        Id=relationship_id,
        Type=DRAWING_REL_TYPE,
        Target=drawing_rel_target,
    )
    parts[sheet_rels_path] = ET.tostring(rels_root, encoding="utf-8", xml_declaration=True)

    sheet_root = ET.fromstring(parts[sheet_part])
    drawing_tag = f"{{{XML_NS['main']}}}drawing"
    if sheet_root.find(drawing_tag) is None:
        sheet_root.append(
            ET.Element(
                drawing_tag,
                {f"{{{XML_NS['office_rel']}}}id": relationship_id},
            )
        )
    parts[sheet_part] = ET.tostring(sheet_root, encoding="utf-8", xml_declaration=True)

    parts[drawing_part] = _drawing_xml(text=text, anchor=anchor, shape_name=shape_name)

    content_types_root = ET.fromstring(parts["[Content_Types].xml"])
    drawing_part_name = "/" + drawing_part
    has_override = any(
        override.attrib.get("PartName") == drawing_part_name
        for override in content_types_root.findall("{http://schemas.openxmlformats.org/package/2006/content-types}Override")
    )
    if not has_override:
        ET.SubElement(
            content_types_root,
            "{http://schemas.openxmlformats.org/package/2006/content-types}Override",
            PartName=drawing_part_name,
            ContentType=DRAWING_CONTENT_TYPE,
        )
        parts["[Content_Types].xml"] = ET.tostring(
            content_types_root,
            encoding="utf-8",
            xml_declaration=True,
        )

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
        tmp_path = Path(tmp_file.name)

    try:
        with ZipFile(tmp_path, "w", compression=ZIP_DEFLATED) as zip_file:
            for name, data in parts.items():
                zip_file.writestr(name, data)
        tmp_path.replace(path)
    finally:
        if tmp_path.exists():
            tmp_path.unlink()
