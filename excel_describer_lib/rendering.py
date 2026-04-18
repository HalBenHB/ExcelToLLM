from __future__ import annotations

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

from .progress import ProgressFn, noop_progress


def _sanitize_cell(value: str) -> str:
    value = value.replace("\r\n", "<br>").replace("\r", "<br>").replace("\n", "<br>")
    value = value.replace("|", "\\|")
    return value


def tabularize_sheet(
    ws,
    xl_pd: pd.ExcelFile,
    sheet: str,
    is_xls: bool,
    source_path=None,
    progress_fn: ProgressFn = noop_progress,
) -> str:
    col_letters_fn = lambda n: [get_column_letter(c) for c in range(1, n + 1)]

    if is_xls:
        df = xl_pd.parse(sheet, header=None)
        if df.empty:
            return "_[empty sheet]_\n"
        max_row, max_col = df.shape
        col_letters = col_letters_fn(max_col)
        lines = []
        lines.append("| Row | " + " | ".join(col_letters) + " |")
        lines.append("|:---:|" + "|".join([":-----"] * max_col) + "|")
        for r_idx in range(max_row):
            cells = [str(r_idx + 1)]
            for c_idx in range(max_col):
                val = df.iat[r_idx, c_idx]
                raw = "" if pd.isna(val) else str(val)
                cells.append(_sanitize_cell(raw))
            lines.append("| " + " | ".join(cells) + " |")
            if (r_idx + 1) % 50 == 0 or r_idx + 1 == max_row:
                progress_fn(r_idx + 1, max_row, f"Rendering row {r_idx + 1}/{max_row}")
        return "\n".join(lines) + "\n"

    formula_map: dict[tuple[int, int], str] = {}
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_map[(cell.row, cell.column)] = cell.value

    workbook_path = source_path or getattr(ws.parent, "path", None)
    if workbook_path is None:
        raise ValueError("A workbook source path is required for .xlsx tabularization.")

    wb_values = openpyxl.load_workbook(workbook_path, data_only=True)
    ws_values = wb_values[sheet]

    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    if max_row == 0 or max_col == 0:
        return "_[empty sheet]_\n"

    col_letters = col_letters_fn(max_col)
    lines = []
    lines.append("| Row | " + " | ".join(col_letters) + " |")
    lines.append("|:---:|" + "|".join([":-----"] * max_col) + "|")

    for row_idx in range(1, max_row + 1):
        cells = [str(row_idx)]
        for col_idx in range(1, max_col + 1):
            val = ws_values.cell(row=row_idx, column=col_idx).value
            formula = formula_map.get((row_idx, col_idx))
            if formula and val is not None:
                raw = f"{val} `{formula}`"
            elif formula:
                raw = f"`{formula}`"
            elif val is None:
                raw = ""
            else:
                raw = str(val)
            cells.append(_sanitize_cell(raw))
        lines.append("| " + " | ".join(cells) + " |")
        if row_idx % 50 == 0 or row_idx == max_row:
            progress_fn(row_idx, max_row, f"Rendering row {row_idx}/{max_row}")

    return "\n".join(lines) + "\n"
