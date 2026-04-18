from __future__ import annotations

import pandas as pd
from openpyxl.utils import get_column_letter

from .constants import PIVOT_AGG_PREFIXES, PIVOT_FILTER_KEYWORDS
from .progress import ProgressFn, noop_progress


def _looks_like_pivot(df_raw: pd.DataFrame, header_row_idx: int) -> bool:
    if header_row_idx == 0:
        return False

    for row_i in range(header_row_idx):
        row_vals = df_raw.iloc[row_i].dropna().astype(str).str.lower().tolist()
        if any(keyword in value for value in row_vals for keyword in PIVOT_FILTER_KEYWORDS):
            return True

    if header_row_idx < len(df_raw):
        header_vals = df_raw.iloc[header_row_idx].dropna().astype(str).str.lower().tolist()
        if any(value.startswith(prefix) for value in header_vals for prefix in PIVOT_AGG_PREFIXES):
            return True

    return False


def find_table_start(df_raw: pd.DataFrame, min_fill_ratio: float = 0.5) -> int:
    candidates: list[int] = []
    for i, row in df_raw.iterrows():
        non_null = row.notna().sum()
        total = len(row)
        if non_null >= 2 and (non_null / total) >= min_fill_ratio:
            candidates.append(i)

    if not candidates:
        return 0

    for a, b in zip(candidates, candidates[1:]):
        rows_between = range(a + 1, b)
        if any(df_raw.iloc[r].notna().sum() <= 1 for r in rows_between):
            return b

    return candidates[0]


def get_formula_cells(ws) -> dict[tuple[int, int], str]:
    if ws is None:
        return {}

    formulas: dict[tuple[int, int], str] = {}
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formulas[(cell.row - 1, cell.column - 1)] = cell.value
    return formulas


def _format_skipped_row(
    df_raw: pd.DataFrame,
    row_i: int,
    formula_cells: dict[tuple[int, int], str],
) -> list[str]:
    out: list[str] = []
    row_data = df_raw.iloc[row_i]
    non_null = row_data.dropna()

    if non_null.empty:
        out.append(f">  **Row {row_i + 1}:** _(blank)_")
        return out

    parts: list[str] = []
    for col_i, val in non_null.items():
        col_letter = get_column_letter(int(col_i) + 1)
        formula = formula_cells.get((row_i, int(col_i)))
        if formula:
            parts.append(f"`{col_letter}`: {val} `{formula}`")
        else:
            parts.append(f"`{col_letter}`: **{val}**")

    out.append(f">  **Row {row_i + 1}:** " + " · ".join(parts))
    return out


def describe_sheet(
    xl_pd: pd.ExcelFile,
    ws,
    sheet: str,
    is_xls: bool,
    floating_text_items: list[dict[str, str]] | None = None,
    manual_header_row_idx: int | None = None,
    max_unique_display: int = 10,
    progress_fn: ProgressFn = noop_progress,
) -> str:
    lines: list[str] = []

    if is_xls:
        lines.append("> ℹ️ **Legacy `.xls` format** — formula introspection is unavailable.")
        lines.append("")

    if floating_text_items:
        lines.append(f"> 🧷 **Floating text objects detected:** {len(floating_text_items)}")
        for item in floating_text_items:
            text = item["text"].replace("\r\n", "\n").replace("\r", "\n")
            text = " / ".join(part.strip() for part in text.split("\n") if part.strip())
            lines.append(f"> - `{item['name']}` near **{item['anchor']}**: {text}")
        lines.append("")

    df_raw = xl_pd.parse(sheet, header=None)

    if df_raw.empty:
        lines.append("_[empty sheet grid — skipped]_")
        return "\n".join(lines)

    if manual_header_row_idx is not None:
        header_row_idx = manual_header_row_idx
        header_source = f"manual override (Excel row {header_row_idx + 1})"
    else:
        header_row_idx = find_table_start(df_raw)
        header_source = "auto-detected"

    if header_row_idx >= len(df_raw):
        header_row_idx = max(len(df_raw) - 1, 0)
        header_source += " - clamped to last non-empty sheet row"

    formula_cells = get_formula_cells(ws)

    if header_row_idx > 0:
        is_pivot = manual_header_row_idx is None and _looks_like_pivot(df_raw, header_row_idx)
        if manual_header_row_idx is not None:
            lines.append(
                f"> 🛠️ **Manual header override applied.** "
                f"Excel row **{header_row_idx + 1}** is treated as the header."
            )
        elif is_pivot:
            lines.append(
                "> 🔄 **Pivot table export detected.**  "
                "The rows above the header contain report-filter selections "
                "and/or grand totals."
            )
        else:
            lines.append(
                f"> ⚠️ Skipped **{header_row_idx}** leading row(s) — "
                "likely notes or fast-calc rows."
            )
        lines.append(">")
        for row_i in range(header_row_idx):
            lines.extend(_format_skipped_row(df_raw, row_i, formula_cells))
        lines.append("")

    df_full = xl_pd.parse(sheet, header=header_row_idx)

    formula_columns = {
        col_idx for (row_idx, col_idx), _formula in formula_cells.items()
        if row_idx >= header_row_idx + 1
    }
    surviving_mask = ~df_full.isna().all(axis=0)
    surviving_positions = [
        idx for idx, keep in enumerate(surviving_mask.tolist())
        if keep or idx in formula_columns
    ]

    df = df_full.iloc[:, surviving_positions].dropna(axis=0, how="all")
    df.columns = [
        f"Unnamed_col_{i}" if str(col).startswith("Unnamed:") else col
        for i, col in enumerate(df.columns)
    ]

    lines.append(f"- **Table starts at:** Excel row {header_row_idx + 1}")
    lines.append(f"- **Header detection:** {header_source}")
    lines.append(f"- **Shape:** {df.shape[0]} rows × {df.shape[1]} cols")
    lines.append("")

    data_start_row = header_row_idx + 1
    total_cols = len(df.columns)

    for col_idx, (col, orig_col_i) in enumerate(zip(df.columns, surviving_positions), 1):
        progress_fn(col_idx, total_cols, f"Analysing column: {str(col)[:35]}")

        col_letter = get_column_letter(orig_col_i + 1)
        nulls = df[col].isna().sum()
        unique = df[col].nunique()
        null_pct = f"{nulls / len(df) * 100:.1f}%" if len(df) > 0 else "n/a"

        col_formulas = [
            formula
            for (row_i, col_i), formula in formula_cells.items()
            if col_i == orig_col_i and row_i >= data_start_row
        ]

        if col_formulas:
            formula_note = f"\n  - ↳ **formula-driven** — sample: `{col_formulas[0]}`"
            if len(set(col_formulas)) > 1:
                formula_note += f" _{len(set(col_formulas))} distinct formulas in column_"
        else:
            formula_note = ""

        unique_vals = df[col].dropna().unique().tolist()
        unique_detail = (
            f"\n  - **All values:** `{unique_vals}`" if unique <= max_unique_display else ""
        )

        col_str = str(col).lower()
        is_agg_col = any(col_str.startswith(prefix) for prefix in PIVOT_AGG_PREFIXES)
        agg_badge = " _(pivot aggregation)_" if is_agg_col else ""

        lines.append(f"#### `{col_letter}` — {col}{agg_badge}")

        if pd.api.types.is_numeric_dtype(df[col]):
            non_null = df[col].dropna()
            if len(non_null) == 0:
                lines.append("- **Type:** numeric | _all nulls_")
            else:
                lines.append(
                    f"- **Type:** numeric | "
                    f"min `{non_null.min():.2f}` · max `{non_null.max():.2f}` · "
                    f"mean `{non_null.mean():.2f}` · median `{non_null.median():.2f}` · "
                    f"total `{non_null.sum():.2f}`"
                )
                lines.append(f"- **Nulls:** {nulls} ({null_pct})")
        elif pd.api.types.is_datetime64_any_dtype(df[col]):
            lines.append(f"- **Type:** date | range `{df[col].min()}` → `{df[col].max()}`")
            lines.append(f"- **Nulls:** {nulls} ({null_pct})")
        else:
            sample_vals = df[col].dropna().unique()[:3].tolist()
            lines.append(f"- **Type:** text | {unique} unique values | sample: `{sample_vals}`")
            lines.append(f"- **Nulls:** {nulls} ({null_pct})")

        if unique_detail:
            lines.append(unique_detail.strip())
        if formula_note:
            lines.append(formula_note.strip())
        lines.append("")

    return "\n".join(lines)
