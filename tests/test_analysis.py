from __future__ import annotations

import unittest
from pathlib import Path

import openpyxl
import pandas as pd

from excel_describer_lib.analysis import describe_sheet, find_table_start
from excel_describer_lib.progress import noop_progress
from excel_describer_lib.test_workbooks import (
    create_formula_workbook,
    create_manual_override_workbook,
    create_pivot_like_workbook,
    create_unnamed_header_workbook,
)
from support import workspace_temp_dir


class AnalysisTests(unittest.TestCase):
    def test_find_table_start_skips_metadata_before_blank_separator(self) -> None:
        df_raw = pd.DataFrame(
            [
                ["Monthly export", None, None],
                [None, None, None],
                ["Region", "Sales", "Margin"],
                ["North", 120, 0.3],
            ]
        )
        self.assertEqual(find_table_start(df_raw), 2)

    def test_describe_sheet_applies_manual_override_and_preserves_metadata(self) -> None:
        with workspace_temp_dir("analysis_manual") as tmp_dir:
            workbook_path = create_manual_override_workbook(tmp_dir / "manual.xlsx")
            xl_pd = pd.ExcelFile(workbook_path)
            workbook = openpyxl.load_workbook(workbook_path, data_only=False)

            description = describe_sheet(
                xl_pd,
                workbook["ManualHeader"],
                "ManualHeader",
                False,
                manual_header_row_idx=2,
                progress_fn=noop_progress,
            )

        self.assertIn("Manual header override applied", description)
        self.assertIn("**Row 1:**", description)
        self.assertIn("**Row 2:**", description)
        self.assertIn("- **Table starts at:** Excel row 3", description)
        self.assertIn("- **Header detection:** manual override (Excel row 3)", description)

    def test_describe_sheet_detects_pivot_like_exports(self) -> None:
        with workspace_temp_dir("analysis_pivot") as tmp_dir:
            workbook_path = create_pivot_like_workbook(tmp_dir / "pivot.xlsx")
            xl_pd = pd.ExcelFile(workbook_path)
            workbook = openpyxl.load_workbook(workbook_path, data_only=False)

            description = describe_sheet(
                xl_pd,
                workbook["PivotLike"],
                "PivotLike",
                False,
                progress_fn=noop_progress,
            )

        self.assertIn("Pivot table export detected", description)
        self.assertIn("Sum of Sales _(pivot aggregation)_", description)
        self.assertIn("- **Table starts at:** Excel row 3", description)

    def test_describe_sheet_reports_formula_driven_columns(self) -> None:
        with workspace_temp_dir("analysis_formula") as tmp_dir:
            workbook_path = create_formula_workbook(tmp_dir / "formula.xlsx")
            xl_pd = pd.ExcelFile(workbook_path)
            workbook = openpyxl.load_workbook(workbook_path, data_only=False)

            description = describe_sheet(
                xl_pd,
                workbook["Formulas"],
                "Formulas",
                False,
                progress_fn=noop_progress,
            )

        self.assertIn("formula-driven", description)
        self.assertIn("=A2*B2", description)

    def test_describe_sheet_normalizes_unnamed_headers(self) -> None:
        with workspace_temp_dir("analysis_unnamed") as tmp_dir:
            workbook_path = create_unnamed_header_workbook(tmp_dir / "unnamed.xlsx")
            xl_pd = pd.ExcelFile(workbook_path)
            workbook = openpyxl.load_workbook(workbook_path, data_only=False)

            description = describe_sheet(
                xl_pd,
                workbook["Unnamed"],
                "Unnamed",
                False,
                progress_fn=noop_progress,
            )

        self.assertIn("Unnamed_col_1", description)


if __name__ == "__main__":
    unittest.main()
