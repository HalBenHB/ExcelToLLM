import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
import re
import sys
from zipfile import ZipFile
import posixpath
import xml.etree.ElementTree as ET

# ── Config ────────────────────────────────────────────────────────────────────
MAX_UNIQUE_DISPLAY = 70
MARKDOWN_STRUCTURE_NOTE = """## How to Read This Markdown

This file is designed for both humans and LLMs.

- Each `## Sheet: ...` section describes one worksheet.
- `Table starts at` tells you which Excel row is treated as the real header row.
- Rows shown in blockquotes above the table description are metadata, notes, filters, fast-calculation rows, or other pre-table content that appeared before the header.
- Floating textboxes and other text-bearing drawing shapes are listed separately when present.
- Each `####` subsection describes one detected column, including its Excel column letter, inferred type, null counts, and sample values.
- If formulas are present, they are shown inline using backticks.
- If `## Table: ...` sections exist near the end, they contain row-by-row Markdown table exports of the selected sheets.

When a sheet uses a manual header override, that row is trusted as the header even if automatic detection would choose a different starting row.
"""

XML_NS = {
    "pkgrel": "http://schemas.openxmlformats.org/package/2006/relationships",
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "office_rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
}


# ── .xls / .xlsx dispatcher ───────────────────────────────────────────────────

def load_workbook_safe(file_path: Path):
    """
    Return (wb, is_xls).
    .xlsx/.xlsm → openpyxl workbook (data_only=False, formulas preserved).
    .xls        → (None, True): openpyxl cannot read these; pandas/xlrd
                  handles them automatically, but formula introspection is
                  unavailable and a notice is written to the output.
    """
    if file_path.suffix.lower() == ".xls":
        return None, True
    wb = openpyxl.load_workbook(file_path, data_only=False)
    wb.path = str(file_path)
    return wb, False


# ── Progress bar ──────────────────────────────────────────────────────────────

def progress_bar(current: int, total: int, label: str = "", width: int = 40) -> None:
    filled = int(width * current / total) if total > 0 else width
    bar = "#" * filled + "-" * (width - filled)
    pct = f"{100 * current / total:.0f}%" if total > 0 else "100%"
    print(f"\r  [{bar}] {pct}  {label:<40}", end="", flush=True)
    if current >= total:
        print()


# ── Interactive prompts ───────────────────────────────────────────────────────

def prompt_file_selection(excel_files: list[Path]) -> Path | None:
    print("\n+-- Excel files found " + "-" * 40)
    for i, f in enumerate(excel_files, 1):
        print(f"|  [{i}] {f.name}")
    print("|  [0] Exit")
    print("+" + "-" * 60)
    while True:
        raw = input("  Select a file (number): ").strip()
        if raw == "0":
            return None
        if raw.isdigit() and 1 <= int(raw) <= len(excel_files):
            return excel_files[int(raw) - 1]
        print(f"  x Enter a number between 0 and {len(excel_files)}.")


# ── Tabularization prompt helpers ─────────────────────────────────────────────

def parse_sheet_selection(raw: str, max_index: int) -> list[int] | None:
    """
    Parse a 1-based sheet selection string.

    Supported inputs:
      - 0            -> skip tabularization
      - a / all      -> select all sheets
      - 1,3,5        -> individual sheet numbers
      - 2-4,7        -> inclusive ranges plus individual picks
    """
    raw = raw.strip().lower()
    if raw in {"0", ""}:
        return []
    if raw in {"a", "all"}:
        return list(range(1, max_index + 1))

    selected: list[int] = []
    seen: set[int] = set()

    for token in re.split(r"[\s,]+", raw):
        if not token:
            continue

        if "-" in token:
            parts = token.split("-", maxsplit=1)
            if len(parts) != 2 or not all(part.isdigit() for part in parts):
                return None

            start, end = map(int, parts)
            if start > end or start < 1 or end > max_index:
                return None

            for idx in range(start, end + 1):
                if idx not in seen:
                    selected.append(idx)
                    seen.add(idx)
            continue

        if not token.isdigit():
            return None

        idx = int(token)
        if idx < 1 or idx > max_index:
            return None
        if idx not in seen:
            selected.append(idx)
            seen.add(idx)

    return selected


def prompt_tabularize(sheet_names: list[str]) -> list[str]:
    print("\n+-- Tabularize sheet tables " + "-" * 33)
    for i, s in enumerate(sheet_names, 1):
        print(f"|  [{i}] {s}")
    print("|  [a] All sheets")
    print("|  [0] Skip - no tabularization")
    print("+" + "-" * 60)
    while True:
        raw = input(
            "  Select sheet(s) to render as Markdown tables "
            "(e.g. 1,3-4 / a / 0): "
        )
        selected = parse_sheet_selection(raw, len(sheet_names))
        if selected is not None:
            return [sheet_names[i - 1] for i in selected]
        print(
            f"  x Enter 0, 'a', or sheet numbers/ranges between 1 and {len(sheet_names)}."
        )


def parse_header_overrides(raw: str, sheet_names: list[str]) -> dict[str, int] | None:
    """
    Parse sheet header overrides in the form `sheet_number:excel_header_row`.

    Example:
      2:3,5:7  -> sheet 2 uses Excel row 3 as header, sheet 5 uses row 7.

    Returns a mapping of sheet name -> 0-based header row index.
    """
    raw = raw.strip().lower()
    if raw in {"", "0"}:
        return {}

    overrides: dict[str, int] = {}
    seen_sheet_numbers: set[int] = set()

    for token in re.split(r"[\s,]+", raw):
        if not token:
            continue

        if ":" not in token:
            return None

        sheet_part, row_part = token.split(":", maxsplit=1)
        if not sheet_part.isdigit() or not row_part.isdigit():
            return None

        sheet_number = int(sheet_part)
        excel_row = int(row_part)

        if sheet_number < 1 or sheet_number > len(sheet_names) or excel_row < 1:
            return None
        if sheet_number in seen_sheet_numbers:
            return None

        seen_sheet_numbers.add(sheet_number)
        overrides[sheet_names[sheet_number - 1]] = excel_row - 1

    return overrides


def prompt_header_overrides(sheet_names: list[str]) -> dict[str, int]:
    print("\n+-- Manual header row overrides " + "-" * 28)
    for i, s in enumerate(sheet_names, 1):
        print(f"|  [{i}] {s}")
    print("|  Format: sheet_number:header_row")
    print("|  Example: 2:3,5:7")
    print("|  [0] No manual overrides")
    print("+" + "-" * 60)
    while True:
        raw = input(
            "  Enter overrides for sheets whose header row is known: "
        )
        overrides = parse_header_overrides(raw, sheet_names)
        if overrides is not None:
            return overrides
        print("  x Enter values like 2:3,5:7 or 0.")


def _normalize_zip_target(base_part: str, target: str) -> str:
    base_dir = posixpath.dirname(base_part)
    normalized = posixpath.normpath(posixpath.join(base_dir, target))
    return normalized.lstrip("/")


def _load_relationships(zip_file: ZipFile, rels_path: str) -> dict[str, tuple[str, str]]:
    if rels_path not in zip_file.namelist():
        return {}

    rels_root = ET.fromstring(zip_file.read(rels_path))
    relationships: dict[str, tuple[str, str]] = {}
    for rel in rels_root.findall("pkgrel:Relationship", XML_NS):
        relationships[rel.attrib["Id"]] = (
            rel.attrib.get("Type", ""),
            rel.attrib.get("Target", ""),
        )
    return relationships


def _extract_shape_text(shape) -> str:
    paragraphs: list[str] = []
    for paragraph in shape.findall("xdr:txBody/a:p", XML_NS):
        fragments = [
            (node.text or "")
            for node in paragraph.findall(".//a:t", XML_NS)
            if (node.text or "").strip()
        ]
        if fragments:
            paragraphs.append("".join(fragments).strip())
    return "\n".join(paragraphs).strip()


def _anchor_to_label(anchor) -> str:
    if anchor is None:
        return "unknown position"

    col_node = anchor.find("xdr:col", XML_NS)
    row_node = anchor.find("xdr:row", XML_NS)
    if col_node is None or row_node is None:
        return "unknown position"

    try:
        col_idx = int(col_node.text or "0")
        row_idx = int(row_node.text or "0")
    except ValueError:
        return "unknown position"

    return f"{get_column_letter(col_idx + 1)}{row_idx + 1}"


def extract_sheet_floating_text(file_path: Path, sheet_names: list[str]) -> dict[str, list[dict[str, str]]]:
    """
    Extract floating textbox-like drawing content from an .xlsx package.

    Only DrawingML shapes with text bodies are included here. Traditional VML
    note/comment containers are ignored because they are comment infrastructure,
    not standalone floating textboxes.
    """
    results = {sheet_name: [] for sheet_name in sheet_names}

    with ZipFile(file_path) as zip_file:
        workbook_root = ET.fromstring(zip_file.read("xl/workbook.xml"))
        workbook_rels = _load_relationships(zip_file, "xl/_rels/workbook.xml.rels")

        sheet_part_by_name: dict[str, str] = {}
        for sheet in workbook_root.find("main:sheets", XML_NS):
            sheet_name = sheet.attrib.get("name")
            rel_id = sheet.attrib.get(f"{{{XML_NS['office_rel']}}}id")
            if not sheet_name or not rel_id or rel_id not in workbook_rels:
                continue

            _, target = workbook_rels[rel_id]
            sheet_part_by_name[sheet_name] = _normalize_zip_target("xl/workbook.xml", target)

        for sheet_name, sheet_part in sheet_part_by_name.items():
            if sheet_name not in results:
                continue

            sheet_filename = posixpath.basename(sheet_part)
            sheet_rels_path = f"xl/worksheets/_rels/{sheet_filename}.rels"
            sheet_rels = _load_relationships(zip_file, sheet_rels_path)

            drawing_targets = [
                target for rel_type, target in sheet_rels.values()
                if rel_type.endswith("/drawing")
            ]

            for drawing_target in drawing_targets:
                drawing_part = _normalize_zip_target(sheet_part, drawing_target)
                if drawing_part not in zip_file.namelist():
                    continue

                drawing_root = ET.fromstring(zip_file.read(drawing_part))
                for anchor_tag in ("twoCellAnchor", "oneCellAnchor", "absoluteAnchor"):
                    for anchor in drawing_root.findall(f"xdr:{anchor_tag}", XML_NS):
                        for shape in anchor.findall("xdr:sp", XML_NS):
                            text = _extract_shape_text(shape)
                            if not text:
                                continue

                            c_nv_pr = shape.find("xdr:nvSpPr/xdr:cNvPr", XML_NS)
                            shape_name = (
                                c_nv_pr.attrib.get("name", "Unnamed shape")
                                if c_nv_pr is not None else "Unnamed shape"
                            )
                            from_anchor = anchor.find("xdr:from", XML_NS)
                            results[sheet_name].append(
                                {
                                    "name": shape_name,
                                    "anchor": _anchor_to_label(from_anchor),
                                    "text": text,
                                }
                            )

    return results



# ── Cell sanitiser ────────────────────────────────────────────────────────────
def _sanitize_cell(value: str) -> str:
    """Collapse newlines to <br> and escape pipes so Markdown table rows stay intact."""
    value = value.replace("\r\n", "<br>").replace("\r", "<br>").replace("\n", "<br>")
    value = value.replace("|", "\\|")
    return value


# ── Sheet tabularizer ─────────────────────────────────────────────────────────

def tabularize_sheet(ws, xl_pd: pd.ExcelFile, sheet: str, is_xls: bool) -> str:
    """
    Render a worksheet as a Markdown table.
    ws=None (is_xls=True): values only via pandas/xlrd — no formula column.
    ws set  (is_xls=False): values from a data_only openpyxl load + formulas
                            from the formula-preserving ws passed in.
    """
    col_letters_fn = lambda n: [get_column_letter(c) for c in range(1, n + 1)]

    # ── .xls path (values only, via pandas) ──────────────────────────────
    if is_xls:
        df = xl_pd.parse(sheet, header=None)
        if df.empty:
            return "_[empty sheet]_\n"
        max_row, max_col = df.shape
        col_letters = col_letters_fn(max_col)
        lines = []
        lines.append("| Row | " + " | ".join(col_letters) + " |")
        lines.append("|:---:|" + "|".join([":-----"] * max_col) + "|")
        for r_idx in range(max_row):
            cells = [str(r_idx + 1)]
            for c_idx in range(max_col):
                val = df.iat[r_idx, c_idx]
                raw = "" if pd.isna(val) else str(val)
                cells.append(_sanitize_cell(raw))
            lines.append("| " + " | ".join(cells) + " |")
            if (r_idx + 1) % 50 == 0 or r_idx + 1 == max_row:
                progress_bar(r_idx + 1, max_row, label=f"Rendering row {r_idx + 1}/{max_row}")
        return "\n".join(lines) + "\n"

    # ── .xlsx path (values + formulas, via openpyxl) ──────────────────────
    formula_map: dict[tuple[int, int], str] = {}
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_map[(cell.row, cell.column)] = cell.value

    wb_values = openpyxl.load_workbook(ws.parent.path, data_only=True)
    ws_values = wb_values[sheet]

    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    if max_row == 0 or max_col == 0:
        return "_[empty sheet]_\n"

    col_letters = col_letters_fn(max_col)
    lines = []
    lines.append("| Row | " + " | ".join(col_letters) + " |")
    lines.append("|:---:|" + "|".join([":-----"] * max_col) + "|")

    for r in range(1, max_row + 1):
        cells = [str(r)]
        for c in range(1, max_col + 1):
            val = ws_values.cell(row=r, column=c).value
            formula = formula_map.get((r, c))
            if formula and val is not None:
                raw = f"{val} `{formula}`"
            elif formula:
                raw = f"`{formula}`"
            elif val is None:
                raw = ""
            else:
                raw = str(val)
            cells.append(_sanitize_cell(raw))
        lines.append("| " + " | ".join(cells) + " |")
        if r % 50 == 0 or r == max_row:
            progress_bar(r, max_row, label=f"Rendering row {r}/{max_row}")

    return "\n".join(lines) + "\n"


# ── Core analysis helpers ─────────────────────────────────────────────────────

# Pivot-table column header keywords (case-insensitive prefix match)
_PIVOT_AGG_PREFIXES = ("sum of ", "count of ", "average of ", "avg of ",
                       "max of ", "min of ", "product of ", "stddev of ")

_PIVOT_FILTER_KEYWORDS = ("(multiple items)", "(all)", "(blank)")


def _looks_like_pivot(df_raw: pd.DataFrame, header_row_idx: int) -> bool:
    """
    Heuristic checks for a pivot table export:
      - There are leading rows before the real header  (filter / grand-total rows)
      - At least one of those leading rows contains a pivot filter sentinel like
        "(Multiple Items)" or "(All)"
      - OR at least one column header starts with "Sum of", "Count of", etc.
    """
    if header_row_idx == 0:
        return False

    # Check leading rows for filter sentinels
    for row_i in range(header_row_idx):
        row_vals = df_raw.iloc[row_i].dropna().astype(str).str.lower().tolist()
        if any(kw in v for v in row_vals for kw in _PIVOT_FILTER_KEYWORDS):
            return True

    # Check header row for aggregation prefixes
    if header_row_idx < len(df_raw):
        header_vals = df_raw.iloc[header_row_idx].dropna().astype(str).str.lower().tolist()
        if any(v.startswith(pfx) for v in header_vals for pfx in _PIVOT_AGG_PREFIXES):
            return True

    return False


def find_table_start(df_raw: pd.DataFrame, min_fill_ratio: float = 0.5) -> int:
    """
    Return the 0-based row index of the table header.

    Strategy
    --------
    1. Collect candidate rows: ≥2 non-null cells AND ≥ min_fill_ratio filled.
    2. Walk consecutive candidate pairs.  If a blank/near-blank row (≤1 non-null
       cell) sits between them, the pair represents a "metadata block → blank →
       real header" pattern common in pivot table exports.  Skip to the candidate
       after the gap.
    3. Fall back to the very first candidate.
    """
    candidates: list[int] = []
    for i, row in df_raw.iterrows():
        non_null = row.notna().sum()
        total = len(row)
        if non_null >= 2 and (non_null / total) >= min_fill_ratio:
            candidates.append(i)

    if not candidates:
        return 0

    for a, b in zip(candidates, candidates[1:]):
        rows_between = range(a + 1, b)
        if any(df_raw.iloc[r].notna().sum() <= 1 for r in rows_between):
            # Real header follows the blank separator
            return b

    return candidates[0]


def get_formula_cells(ws) -> dict:
    """Returns formula map, or empty dict when ws is None (legacy .xls)."""
    if ws is None:
        return {}
    formulas = {}
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formulas[(cell.row - 1, cell.column - 1)] = cell.value
    return formulas


def _format_skipped_row(df_raw: pd.DataFrame, row_i: int,
                        formula_cells: dict, ws) -> list[str]:
    """
    Render one skipped leading row as blockquote lines:
      - Blank rows get a _(blank)_ marker.
      - Non-blank rows show each non-null cell as  `ColLetter: value`
        and append any formula found in that row.
    """
    out: list[str] = []
    row_data = df_raw.iloc[row_i]
    non_null = row_data.dropna()

    if non_null.empty:
        out.append(f">  **Row {row_i + 1}:** _(blank)_")
        return out

    parts: list[str] = []
    for col_i, val in non_null.items():
        col_letter = get_column_letter(int(col_i) + 1)
        formula = formula_cells.get((row_i, int(col_i)))
        if formula:
            parts.append(f"`{col_letter}`: {val} `{formula}`")
        else:
            parts.append(f"`{col_letter}`: **{val}**")

    out.append(f">  **Row {row_i + 1}:** " + " · ".join(parts))
    return out


def describe_sheet(
    xl_pd: pd.ExcelFile,
    ws,
    sheet: str,
    is_xls: bool,
    floating_text_items: list[dict[str, str]] | None = None,
    manual_header_row_idx: int | None = None,
    max_unique_display: int = 10,
) -> str:
    lines = []

    if is_xls:
        lines.append("> ℹ️ **Legacy `.xls` format** — formula introspection is unavailable.")
        lines.append("")

    if floating_text_items:
        lines.append(f"> 🧷 **Floating text objects detected:** {len(floating_text_items)}")
        for item in floating_text_items:
            text = item["text"].replace("\r\n", "\n").replace("\r", "\n")
            text = " / ".join(part.strip() for part in text.split("\n") if part.strip())
            lines.append(
                f"> - `{item['name']}` near **{item['anchor']}**: {text}"
            )
        lines.append("")

    df_raw = xl_pd.parse(sheet, header=None)

    if df_raw.empty:
        lines.append("_[empty sheet grid — skipped]_")
        return "\n".join(lines)

    if manual_header_row_idx is not None:
        header_row_idx = manual_header_row_idx
        header_source = f"manual override (Excel row {header_row_idx + 1})"
    else:
        header_row_idx = find_table_start(df_raw)
        header_source = "auto-detected"

    if header_row_idx >= len(df_raw):
        header_row_idx = max(len(df_raw) - 1, 0)
        header_source += " - clamped to last non-empty sheet row"

    formula_cells = get_formula_cells(ws)

    # ── Pivot / leading-row block ─────────────────────────────────────────
    if header_row_idx > 0:
        is_pivot = manual_header_row_idx is None and _looks_like_pivot(df_raw, header_row_idx)
        if manual_header_row_idx is not None:
            lines.append(
                f"> 🛠️ **Manual header override applied.** "
                f"Excel row **{header_row_idx + 1}** is treated as the header."
            )
        elif is_pivot:
            lines.append(
                "> 🔄 **Pivot table export detected.**  "
                "The rows above the header contain report-filter selections "
                "and/or grand totals."
            )
        else:
            lines.append(
                f"> ⚠️ Skipped **{header_row_idx}** leading row(s) — "
                "likely notes or fast-calc rows."
            )
        lines.append(">")
        for row_i in range(header_row_idx):
            lines.extend(_format_skipped_row(df_raw, row_i, formula_cells, ws))
        lines.append("")

    # ── Re-parse with detected header ────────────────────────────────────
    df_full = xl_pd.parse(sheet, header=header_row_idx)

    surviving_mask = ~df_full.isna().all(axis=0)
    surviving_positions = [i for i, keep in enumerate(surviving_mask) if keep]

    df = df_full.loc[:, surviving_mask].dropna(axis=0, how="all")
    df.columns = [
        f"Unnamed_col_{i}" if str(c).startswith("Unnamed:") else c
        for i, c in enumerate(df.columns)
    ]

    lines.append(f"- **Table starts at:** Excel row {header_row_idx + 1}")
    lines.append(f"- **Header detection:** {header_source}")
    lines.append(f"- **Shape:** {df.shape[0]} rows × {df.shape[1]} cols")
    lines.append("")

    data_start_row = header_row_idx + 1
    total_cols = len(df.columns)

    for col_idx, (col, orig_col_i) in enumerate(zip(df.columns, surviving_positions), 1):
        progress_bar(col_idx, total_cols, label=f"Analysing column: {str(col)[:35]}")

        col_letter = get_column_letter(orig_col_i + 1)
        nulls = df[col].isna().sum()
        unique = df[col].nunique()
        null_pct = f"{nulls / len(df) * 100:.1f}%" if len(df) > 0 else "n/a"

        col_formulas = [
            formula
            for (r, c), formula in formula_cells.items()
            if c == orig_col_i and r >= data_start_row
        ]

        if col_formulas:
            formula_note = f"\n  - ↳ **formula-driven** — sample: `{col_formulas[0]}`"
            if len(set(col_formulas)) > 1:
                formula_note += f" _{len(set(col_formulas))} distinct formulas in column_"
        else:
            formula_note = ""

        unique_vals = df[col].dropna().unique().tolist()
        unique_detail = (
            f"\n  - **All values:** `{unique_vals}`" if unique <= max_unique_display else ""
        )

        # Flag columns that look like pivot aggregations
        col_str = str(col).lower()
        is_agg_col = any(col_str.startswith(pfx) for pfx in _PIVOT_AGG_PREFIXES)
        agg_badge = " _(pivot aggregation)_" if is_agg_col else ""

        lines.append(f"#### `{col_letter}` — {col}{agg_badge}")

        if pd.api.types.is_numeric_dtype(df[col]):
            non_null = df[col].dropna()
            if len(non_null) == 0:
                lines.append("- **Type:** numeric | _all nulls_")
            else:
                lines.append(
                    f"- **Type:** numeric | "
                    f"min `{non_null.min():.2f}` · max `{non_null.max():.2f}` · "
                    f"mean `{non_null.mean():.2f}` · median `{non_null.median():.2f}` · "
                    f"total `{non_null.sum():.2f}`"
                )
                lines.append(f"- **Nulls:** {nulls} ({null_pct})")

        elif pd.api.types.is_datetime64_any_dtype(df[col]):
            lines.append(
                f"- **Type:** date | range `{df[col].min()}` → `{df[col].max()}`"
            )
            lines.append(f"- **Nulls:** {nulls} ({null_pct})")

        else:
            sample_vals = df[col].dropna().unique()[:3].tolist()
            lines.append(
                f"- **Type:** text | {unique} unique values | sample: `{sample_vals}`"
            )
            lines.append(f"- **Nulls:** {nulls} ({null_pct})")

        if unique_detail:
            lines.append(unique_detail.strip())
        if formula_note:
            lines.append(formula_note.strip())
        lines.append("")

    return "\n".join(lines)


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    script_dir = Path(__file__).parent.absolute()
    excel_files = list(script_dir.glob("*.xlsx")) + list(script_dir.glob("*.xls"))

    if not excel_files:
        print(f"No Excel files found in {script_dir}")
        sys.exit(0)

    file_path = prompt_file_selection(excel_files)
    if file_path is None:
        print("  Exiting.")
        sys.exit(0)

    print(f"\n  Processing: {file_path.name}")

    try:
        xl_pd = pd.ExcelFile(file_path)
        wb, is_xls = load_workbook_safe(file_path)

        if is_xls:
            print("  [i] Legacy .xls detected - formula introspection disabled.")

        sheet_names = xl_pd.sheet_names
        floating_text_by_sheet = {} if is_xls else extract_sheet_floating_text(file_path, sheet_names)
        tab_sheets = prompt_tabularize(sheet_names)
        header_overrides = prompt_header_overrides(sheet_names)
        output_path = file_path.with_suffix(".md")

        print(f"\n  Writing -> {output_path.name}")
        total_sheets = len(sheet_names)

        with open(output_path, "w", encoding="utf-8") as f:
            f.write(f"# {file_path.name}\n\n")
            f.write(MARKDOWN_STRUCTURE_NOTE + "\n\n")

            for sheet_idx, sheet in enumerate(sheet_names, 1):
                print(f"\n  Sheet {sheet_idx}/{total_sheets}: '{sheet}'")
                f.write(f"## Sheet: {sheet}\n\n")

                ws = wb[sheet] if wb is not None else None
                description = describe_sheet(
                    xl_pd,
                    ws,
                    sheet,
                    is_xls,
                    floating_text_items=floating_text_by_sheet.get(sheet, []),
                    manual_header_row_idx=header_overrides.get(sheet),
                    max_unique_display=MAX_UNIQUE_DISPLAY,
                )
                f.write(description + "\n\n")

                progress_bar(sheet_idx, total_sheets, label=f"Sheet '{sheet}' done")

            if tab_sheets:
                f.write("---\n\n")
                total_tables = len(tab_sheets)
                for table_idx, tab_sheet in enumerate(tab_sheets, 1):
                    print(
                        f"\n  Rendering table {table_idx}/{total_tables}: "
                        f"'{tab_sheet}' as Markdown table..."
                    )
                    f.write(f"## Table: {tab_sheet}\n\n")
                    ws = wb[tab_sheet] if wb is not None else None
                    md_table = tabularize_sheet(ws, xl_pd, tab_sheet, is_xls)
                    f.write(md_table + "\n")

        print(f"\n  [ok] Written to {output_path.name}")

    except Exception as e:
        print(f"\n  [x] Error processing {file_path.name}: {e}")
        raise


if __name__ == "__main__":
    main()
